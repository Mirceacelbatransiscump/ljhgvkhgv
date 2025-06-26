import pandas as pd
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --- Load and clean input files ---
df_demand = pd.read_csv('Customer request.csv', index_col=0)
df_demand.columns = df_demand.columns.str.strip().str.replace('\ufeff|\r|\n', '', regex=True)
df_tree = pd.read_csv('tree_of_operations.csv')  # <- use your new tree here!
df_workers = pd.read_csv('worket shifts and hourly changes.csv', encoding='latin1')
df_starting_stock = pd.read_csv('StartingStock.csv')

shift_hours = 7.5
days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
weeks = [col for col in df_demand.columns if col.lower().startswith('wk')]
num_weeks = min(5, len(weeks))
shifts = sorted(df_workers['Shift'].astype(str).unique())

def operation_sorter(value):
    if isinstance(value, str) and value.strip().lower() in ["final step", "finalstep"]:
        return (9999,)
    try:
        return (int(value),)
    except Exception:
        return (10000, str(value))

def get_machine_limit(row):
    # For future extension: use Workers per machine from tree
    return int(row.get('Workers per machine', 1))

def get_starting_stock(project, machine):
    match = df_starting_stock[
        (df_starting_stock['Project'].str.strip().str.upper() == str(project).strip().upper()) &
        (df_starting_stock['Machine'].str.strip().str.upper() == str(machine).strip().upper())
    ]
    if not match.empty:
        return float(match.iloc[0]['Starting stock'])
    return 0

# Prepare operator pool and shift lookup
operator_pool = list(df_workers['Name'] + ' ' + df_workers['Surname'])
operator_shift = {
    f"{row['Name']} {row['Surname']}": str(row['Shift'])
    for _, row in df_workers.iterrows()
}

wb = openpyxl.Workbook()
del wb['Sheet']

for week_num in range(num_weeks):
    week_name = weeks[week_num]
    ws = wb.create_sheet(f"Week {week_num+1}")
    ws.append([f"Week {week_num+1} ({week_name})"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(days_of_week)*3)
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    week_demand = df_demand[week_name]
    projects = week_demand.index
    steps_per_project = {
        p: df_tree[df_tree['Project'].str.strip().str.upper() == p.strip().upper()]
            .sort_values('orders of operations', key=lambda col: col.map(operation_sorter))
            .reset_index(drop=True)
        for p in projects
    }

    # Build per-step, per-project structure
    step_machine = {}
    step_hourly = {}
    step_limit = {}
    n_steps_per_project = {}
    for p in projects:
        steps = steps_per_project[p]
        n_steps_per_project[p] = len(steps)
        for s, row in steps.iterrows():
            mname = row['Machine']
            step_machine[(p, s)] = mname
            step_hourly[(p, s)] = float(row['Hourly prod'])
            step_limit[(p, s)] = get_machine_limit(row)

    # 1. Calculate all shifts needed for the week for each step, respecting initial stock
    shifts_needed = {}
    remaining_qty = {}
    total_to_produce_per_project_machine = {}
    initial_stock_per_project_machine = {}
    for p in projects:
        steps = steps_per_project[p]
        demand = float(week_demand.loc[p])
        for s, row in steps.iterrows():
            mname = row['Machine']
            hourly = float(row['Hourly prod'])
            initial = get_starting_stock(p, mname)
            if s == 0:
                to_produce = max(demand - initial, 0)
            else:
                to_produce = max(demand, 0)
            shifts_required = int(math.ceil(to_produce / (hourly * shift_hours))) if hourly > 0 else 0
            shifts_needed[(p, s)] = shifts_required
            remaining_qty[(p, s)] = to_produce
            key = (p, mname)
            total_to_produce_per_project_machine[key] = to_produce + initial
            initial_stock_per_project_machine[key] = initial

    # 2. Assign shifts across the week, day-by-day, shift-by-shift with the "C" constraint
    operator_day_shift_assignments = {op: [] for op in operator_pool}
    shifts_assigned = {(p, s): 0 for (p, s) in shifts_needed}
    slots = [(day, shift) for day in days_of_week for shift in shifts]

    # For progress tracking:
    produced_by_day = {}  # (project, machine, day): cumulative produced till end of day

    for day in days_of_week:
        machines_used_1_2 = set()
        machines_used_C = set()
        # Track per-step C assignments for multiworker
        shift_C_assignments = {(p, s): 0 for (p, s) in shifts_needed}
        for shift in shifts:
            ops_available = [
                op for op in operator_pool
                if operator_shift[op] == shift and len(operator_day_shift_assignments[op]) < len(days_of_week)*len(shifts)
            ]
            ops_queue = ops_available.copy()

            # For shift C, prioritize multiworker machines
            if shift == "C":
                steps_to_assign = sorted(
                    [(p, s) for (p, s), needed in shifts_needed.items() if shifts_assigned[(p, s)] < needed],
                    key=lambda tup: -step_limit[tup]  # Sort by max allowed operators, descending
                )
            else:
                steps_to_assign = [
                    (p, s) for (p, s), needed in shifts_needed.items() if shifts_assigned[(p, s)] < needed
                ]

            for (p, s) in steps_to_assign:
                needed = shifts_needed[(p, s)]
                if shifts_assigned[(p, s)] >= needed:
                    continue
                m = step_machine[(p, s)]
                max_workers = step_limit[(p, s)]

                # MACHINE C CONSTRAINTS
                if shift == "C":
                    if m in machines_used_1_2:
                        continue  # Can't use this machine in shift C if used in shift 1/2
                    # Assign up to max_workers operators to this machine for shift C
                    while (shift_C_assignments[(p, s)] < max_workers and ops_queue and shifts_assigned[(p, s)] < needed):
                        op = ops_queue.pop(0)
                        hourly = step_hourly[(p, s)]
                        max_produce = hourly * shift_hours
                        remaining = remaining_qty[(p, s)]
                        produce_now = min(max_produce, remaining)
                        operator_day_shift_assignments[op].append((day, shift, m, p, s+1, produce_now))
                        shifts_assigned[(p, s)] += 1
                        shift_C_assignments[(p, s)] += 1
                        remaining_qty[(p, s)] -= produce_now
                        machines_used_C.add(m)
                else:
                    if m in machines_used_C:
                        continue  # Can't use this machine in shift 1/2 if used in C
                    if not ops_queue:
                        continue
                    op = ops_queue.pop(0)
                    hourly = step_hourly[(p, s)]
                    max_produce = hourly * shift_hours
                    remaining = remaining_qty[(p, s)]
                    produce_now = min(max_produce, remaining)
                    operator_day_shift_assignments[op].append((day, shift, m, p, s+1, produce_now))
                    shifts_assigned[(p, s)] += 1
                    remaining_qty[(p, s)] -= produce_now
                    machines_used_1_2.add(m)

            # Fill in 'Assente' for unassigned operators as before
            for op in ops_available:
                already_assigned = [
                    x for x in operator_day_shift_assignments[op]
                    if x[0] == day and x[1] == shift
                ]
                if not already_assigned:
                    operator_day_shift_assignments[op].append((day, shift, "Assente", "Assente", "", 0))

    # --- Output as human-readable plan ---
    # Italian-style readable table, 2 header rows
    header_top = [''] * 3
    for day in days_of_week:
        header_top.extend([day] * 3)
    header_mid = ['Nome Op', 'Turno', 'Ore']
    for day in days_of_week:
        header_mid += ['Postazione', 'Modello', 'PROD. PREVISTA']
    ws.append(header_top)
    ws.append(header_mid)

    for op in operator_pool:
        shift_val = operator_shift.get(op, "")
        shift_label = f"{shift_val}°" if shift_val.isdigit() else shift_val
        row = [op, shift_label, shift_hours]
        plan_dict = {(entry[0], str(entry[1])): entry for entry in operator_day_shift_assignments[op]}
        for day in days_of_week:
            tup = None
            for shift in shifts:
                assignment = plan_dict.get((day, shift), None)
                if assignment and assignment[2] != "Assente":
                    tup = assignment
                    break
            if tup and tup[2] != "Assente":
                machine = tup[2]
                modello = tup[3]
                prod = int(float(tup[5])) if tup[5] else 0
                row += [machine, modello, prod]
            else:
                row += ['Assente', 'Assente', 0]
        ws.append(row)
        for col in range(1, len(row)+1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if "assente" in str(cell.value).lower():
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            elif col > 3 and str(row[col-1]).strip() and row[col-1] != 0:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # --- PROGRESS BAR SECTION ---
    ws.append([])
    ws.append(['PROGRESSO SETTIMANALE (per progetto/macchina):'])
    # Header for progress
    progress_header = ['Progetto', 'Macchina']
    for day in days_of_week:
        progress_header.append(f"{day}: % completato")
    progress_header.append("Pronto entro venerdì?")
    ws.append(progress_header)

    # --- Calculate produced_by_day for all (project, machine, day) ---
    produced_by_day = {}
    for op in operator_pool:
        for entry in operator_day_shift_assignments[op]:
            day, shift, m, p, step, produced = entry
            if m in ["Assente", "Support"] or not isinstance(produced, (int, float)):
                continue
            key = (p, m, day)
            produced_by_day[key] = produced_by_day.get(key, 0) + float(produced)

    for (p, m) in total_to_produce_per_project_machine:
        initial_stock = initial_stock_per_project_machine[(p, m)]
        total_to_produce = total_to_produce_per_project_machine[(p, m)]
        daily_cumulative = []
        cumulative = initial_stock
        ready = False
        for i, day in enumerate(days_of_week):
            produced = produced_by_day.get((p, m, day), 0)
            cumulative += produced
            percent = 100 * cumulative / total_to_produce if total_to_produce > 0 else 0
            if percent >= 100 and not ready:
                ready = True
            daily_cumulative.append(min(percent, 140))  # cap at 140%
        ready_text = "Sì" if daily_cumulative[-1] >= 100 else "NO"
        row = [p, m] + [f"{v:.0f}%" for v in daily_cumulative] + [ready_text]
        ws.append(row)
        for col in range(3, 3+len(days_of_week)):
            cell = ws.cell(row=ws.max_row, column=col)
            val = daily_cumulative[col-3]
            if val >= 100:
                cell.fill = PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid")
            elif val >= 90:
                cell.fill = PatternFill(start_color="FFF699", end_color="FFF699", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
        ready_cell = ws.cell(row=ws.max_row, column=3+len(days_of_week))
        if ready_text == "Sì":
            ready_cell.fill = PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid")
        else:
            ready_cell.fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")

    # Auto column width
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max(max_length + 2, 12)
        ws.column_dimensions[column_letter].width = adjusted_width

wb.save("dynamic_parallel_flow_with_starting_stock_human_readable_progress.xlsx")
print("Exported human-readable schedule and weekly progress to dynamic_parallel_flow_with_starting_stock_human_readable_progress.xlsx")
